# -*- coding: utf-8 -*-
"""kaoping_core 纯函数自动化测试（无需第三方测试框架，python tests/test_core.py 即可运行）"""
import os
import shutil
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import kaoping_core as kc  # noqa: E402


def test_build_filename():
    # 默认模板
    assert kc.build_filename(kc.DEFAULT_NAME_PATTERN, 2026, 8, "张三") == \
        "班组长安全生产责任制履职清单考评表(8月张三).doc"
    # 全角/自定义模板
    assert kc.build_filename("{XXX}（{Y}年{X}月）", 2026, 12, "李四") == \
        "李四（2026年12月）.doc"
    # 非法字符替换
    assert kc.build_filename("{XXX}", 2026, 8, '王五/李六*') == "王五_李六_.doc"
    # Windows 保留设备名
    assert kc.build_filename("CON", 2026, 8, "x").startswith("_")
    assert kc.build_filename("NUL.{XXX}", 2026, 8, "x").startswith("_")
    # 结尾点/空格消毒
    assert kc.build_filename("{XXX}.doc", 2026, 8, "赵六. ") == "赵六.doc"


def test_replace_slot():
    body = "考评对象（班组长）：  孙忠      管理者姓名：x评价月份：                 评价人员签字"
    b = kc._replace_slot(body, "考评对象（班组长）：", "管理者姓名", "测试员")
    assert "：  测试员      管理者姓名" in b
    b = kc._replace_slot(b, "评价月份：", "评价人员签字", "8月")
    seg = b[b.find("评价月份：") + 5:b.find("评价人员签字")]
    assert seg.strip() == "8月", f"月份区间异常: {seg!r}"
    # 值居中：前后都有空白
    assert seg.startswith(" ") and seg.endswith(" ")


def test_item_rows():
    """行映射：第 1 项为双行(R3,R4)，其余每项一行。"""
    assert kc.N_ITEMS == 17
    assert kc._item_rows(1) == [3, 4]
    assert kc._item_rows(2) == [5]
    assert kc._item_rows(17) == [20]
    assert kc.TOTAL_ROW == 21


def test_sum_scores():
    items = {
        1: {"score": "10", "super_score": "10", "sub": {"score": "8", "super_score": "9"}},
        2: {"score": "5", "super_score": ""},
        3: {"score": "abc", "super_score": "5"},
    }
    # 第 1 项两行得分 10+8、第 2 项 5、第 3 项 abc 忽略 → 23
    assert kc._sum_scores(items, "score") == "23"
    # 上级：10+9+5 = 24
    assert kc._sum_scores(items, "super_score") == "24"
    assert kc._sum_scores({}, "score") == ""


def test_is_image():
    assert kc.is_image("a.JPG")
    assert kc.is_image("a.png")
    assert not kc.is_image("a.docx")
    assert not kc.is_image("a.xlsx")


def test_match_materials_file():
    assert kc.match_materials_file("扫雷/2026年扫雷统计表.xlsx") == 10
    assert kc.match_materials_file("综合大检查/安全综合检查通报.doc") == 7
    assert kc.match_materials_file("（炼铁）8月主题工作/1.标准化作业/xxx.jpg") == 16
    assert kc.match_materials_file("应急演练/皮带机演练方案.doc") == 8
    assert kc.match_materials_file("班前活动/班前喊话记录.doc") == 9
    assert kc.match_materials_file("安全培训课件.pdf") == 5
    assert kc.match_materials_file("违章查处记录.xls") == 3
    assert kc.match_materials_file("防火防爆/消防器材点检表.xlsx") == 14
    # 无匹配
    assert kc.match_materials_file("高温天气预警/8.3.jpg") is None
    assert kc.match_materials_file("领导视察/接待安排.docx") is None


def test_scan_materials_folder():
    tmp = tempfile.mkdtemp()
    try:
        os.makedirs(os.path.join(tmp, "扫雷"))
        os.makedirs(os.path.join(tmp, "应急演练"))
        # 匹配文件
        open(os.path.join(tmp, "扫雷", "扫雷统计表.xlsx"), "w").close()
        open(os.path.join(tmp, "应急演练", "演练方案.doc"), "w").close()
        # 不匹配文件
        open(os.path.join(tmp, "高温预警.jpg"), "w").close()
        # 应被跳过的考评表文件
        open(os.path.join(tmp, "班组长安全生产责任制履职清单考评表（张三8月）.doc"), "w").close()
        # 应被跳过的临时文件
        open(os.path.join(tmp, "~$临时.docx"), "w").close()

        result, unmatched = kc.scan_materials_folder(tmp)
        assert len(result[10]) == 1 and result[10][0].endswith("扫雷统计表.xlsx")
        assert len(result[8]) == 1 and result[8][0].endswith("演练方案.doc")
        assert len(result[1]) == 0
        assert len(unmatched) == 1, f"未匹配应为1个，实际 {len(unmatched)}"
        assert unmatched[0].endswith("高温预警.jpg")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_fit_size():
    # 已在 2cm 内（含浮点尾差）原样返回——绝不能进入缩放分支，
    # 否则再次 ScaleWidth=100 会把已缩好的对象放大回原始大小（曾实测 486×864pt）。
    assert kc._fit_size(31.9, 56.700001, 2.0) == (31.9, 56.700001)
    assert kc._fit_size(56.7, 56.7, 2.0) == (56.7, 56.7)
    assert kc._fit_size(56.4, 42.0, 2.0) == (56.4, 42.0)
    # 超限等比缩小，长边 ≤ 2cm（56.7pt）
    tw, th = kc._fit_size(486, 864, 2.0)
    assert tw <= 56.7 + 1e-6 and th <= 56.7 + 1e-6
    assert abs(tw / th - 486 / 864) < 0.01
    tw, th = kc._fit_size(2400, 1800, 2.0)
    assert tw <= 56.7 + 1e-6 and th <= 56.7 + 1e-6
    # 异常输入不抛错
    assert kc._fit_size(0, 100, 2.0) == (0, 100)
    assert kc._fit_size(-5, 100, 2.0) == (-5, 100)


def test_prepare_image_file():
    from PIL import Image
    tmp = tempfile.mkdtemp()
    try:
        big = os.path.join(tmp, "big.jpg")
        Image.new("RGB", (3000, 2000)).save(big)
        small = os.path.join(tmp, "small.png")
        Image.new("RGB", (800, 600)).save(small)
        p1, is_tmp1 = kc._prepare_image_file(big, tmp)
        assert is_tmp1 is True
        with Image.open(p1) as im:
            assert max(im.size) <= kc.MAX_EMBED_PX
        p2, is_tmp2 = kc._prepare_image_file(small, tmp)
        assert is_tmp2 is False and p2 == small
        # 非法图片回退原文件，不抛错
        bad = os.path.join(tmp, "bad.jpg")
        with open(bad, "w", encoding="utf-8") as f:
            f.write("not an image")
        p3, is_tmp3 = kc._prepare_image_file(bad, tmp)
        assert p3 == bad and is_tmp3 is False
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _make_fake_lvzhilv():
    """构造一个与真实月度履职履责表同构的临时 xlsx（班组长月度履职评价表）。"""
    import openpyxl
    tmp = tempfile.mkdtemp()
    p = os.path.join(tmp, "履职履责.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "班组长月度履职评价表"
    ws.cell(1, 1, "班组长（事业部）月度安全生产履职评价表")
    ws.cell(2, 1, "序号"); ws.cell(2, 2, "单位"); ws.cell(2, 3, "岗位"); ws.cell(2, 4, "姓名")
    ws.cell(2, 5, "考核维度（100分）"); ws.cell(2, 38, "得分\n排序")
    # 第 1 项 20 分，其余 16 项各 5 分（列 5,7,9,...37）
    names = ["安全绩效", "履职评价", "违章检查", "劳防及工机具", "教育培训", "会议活动",
             "检查改进", "应急处置", "班前活动", "安全隐患", "安全交底", "安全标志",
             "作业许可", "防火防爆", "治安管理", "基础工作", "安全技能"]
    scores = [20] + [5] * 16
    for j, (col, s) in enumerate(zip(range(5, 39, 2), scores)):
        ws.cell(3, col, "%d分" % s)
        ws.cell(4, col, names[j])
        ws.cell(4, col + 1, "扣分说明")
    ws.cell(5, 1, 1); ws.cell(5, 2, "兴达公司"); ws.cell(5, 3, "事业部班组长"); ws.cell(5, 4, "季忠")
    for col, val in zip(range(5, 39, 2), [18] + [5] * 16):
        ws.cell(5, col, val)
    ws.cell(5, 6, "第1项第2行未达标")
    ws.cell(5, 10, "未按违章查处")
    ws.cell(5, 38, 93)
    ws.cell(6, 4, "王旺")
    ws.cell(6, 5, 20); ws.cell(6, 7, 5)
    wb.save(p)
    return tmp, p


def test_read_eval_scores():
    tmp, p = _make_fake_lvzhilv()
    try:
        d = kc.read_eval_scores(p, "季忠")
        assert d["name"] == "季忠"
        assert d["total"] == "93"
        assert d["items"][1]["score"] == "18"
        assert d["items"][3]["score"] == "5"
        assert d["items"][3]["desc"] == ""
        assert d["items"][2]["desc"] == ""
        assert len(d["items"]) == 17
        # 未找到姓名：报错并列出可用姓名
        try:
            kc.read_eval_scores(p, "张三")
            assert False, "应抛出 ValueError"
        except ValueError as e:
            assert "季忠" in str(e) and "王旺" in str(e)
        # 无目标 sheet：报错
        import openpyxl
        bad = os.path.join(tmp, "其他.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = "其他表"
        wb.save(bad)
        try:
            kc.read_eval_scores(bad, "季忠")
            assert False, "应抛出 ValueError"
        except ValueError as e:
            assert "sheet" in str(e)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_list_eval_names():
    tmp, p = _make_fake_lvzhilv()
    try:
        assert kc.list_eval_names(p) == ["季忠", "王旺"]
        # 无目标 sheet 时同样报错
        import openpyxl
        bad = os.path.join(tmp, "其他.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = "其他表"
        wb.save(bad)
        try:
            kc.list_eval_names(bad)
            assert False, "应抛出 ValueError"
        except ValueError as e:
            assert "sheet" in str(e)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_month_from_filename():
    assert kc._month_from_filename("安全员安全生产责任制履职清单考评表(孙忠8月).doc") == "8"
    assert kc._month_from_filename("班组长安全生产责任制履职清单考评表（季忠7月）.doc") == "7"
    assert kc._month_from_filename("模板.doc") == ""


def test_excel_ext():
    assert kc._excel_ext(b"PK\x03\x04xxxx") == ".xlsx"        # xlsx 内容
    assert kc._excel_ext(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1") == ".xls"   # 旧版 xls 内容
    assert kc._excel_ext(b"") == ".xlsx"


def test_parse_kaoping_header():
    hdr = "考评对象（安全员）：孙忠  管理者姓名：（手签）评价月份：7月评价人员签字：（手签）\r\x07"
    name, month = kc._parse_kaoping_header(hdr)
    assert name == "孙忠" and month == "7"
    # 半角冒号 / 缺月份
    name2, month2 = kc._parse_kaoping_header("考评对象（安全员）: 王旺 管理者姓名:")
    assert name2 == "王旺" and month2 == ""
    assert kc._parse_kaoping_header("无关文本") == ("", "")


def test_extract_ole10native():
    # 构造 Ole10Native：4字节长度 + 0200 + GBK文件名 + NUL + 内容长度 + ZIP 内容
    content = b"PK\x03\x04" + b"hello" * 10 + b"PK\x05\x06" + b"\x00\x00"
    body = (b"\x02\x00" + "材料表.xlsx".encode("gbk") + b"\x00"
            + len(content).to_bytes(4, "little") + content)
    raw = len(body).to_bytes(4, "little") + body
    fn, ct = kc._extract_ole10native(raw)
    assert fn == "材料表.xlsx" and ct == content
    # OLE2(.doc) 内容
    ole_content = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 32
    body2 = (b"\x02\x00" + b"a.doc\x00"
             + len(ole_content).to_bytes(4, "little") + ole_content)
    raw2 = len(body2).to_bytes(4, "little") + body2
    fn2, ct2 = kc._extract_ole10native(raw2)
    assert fn2 == "a.doc" and ct2 == ole_content
    # 非法流
    assert kc._extract_ole10native(b"\x00" * 10) == (None, None)


def test_safe_filename():
    assert kc._safe_filename('a/b:c*d?e') == "a_b_c_d_e"
    assert kc._safe_filename("   ") == "未命名"


def test_sniff_office_kind():
    # xlsx：zip 内含 xl/
    import io
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", "")
        z.writestr("xl/workbook.xml", "")
    assert kc._sniff_office_kind(buf.getvalue()) == "excel"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", "")
        z.writestr("word/document.xml", "")
    assert kc._sniff_office_kind(buf.getvalue()) == "word_docx"
    assert kc._sniff_office_kind(b"D0CF11E0A1B11AE1") == "file"
    assert kc._sniff_office_kind(b"") == "file"
    assert kc._sniff_office_kind(None) == "file"


def test_ole_kind_match():
    assert kc._ole_kind_match("excel", "excel") is True
    assert kc._ole_kind_match("excel", "file") is True
    assert kc._ole_kind_match("excel", "pkg") is False
    assert kc._ole_kind_match("word_docx", "pkg") is True
    assert kc._ole_kind_match("pkg", "excel") is True


def main():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failed = 0
    for t in tests:
        try:
            t()
            print(f"[PASS] {t.__name__}")
        except AssertionError as e:
            failed += 1
            print(f"[FAIL] {t.__name__}: {e}")
        except Exception as e:
            failed += 1
            print(f"[ERROR] {t.__name__}: {type(e).__name__}: {e}")
    print(f"\n共 {len(tests)} 项测试，失败 {failed} 项")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
