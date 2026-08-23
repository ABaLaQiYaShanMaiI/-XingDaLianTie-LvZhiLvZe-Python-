# -*- coding: utf-8 -*-
"""kaoping_core 纯函数自动化测试（无需第三方测试框架，python tests/test_core.py 即可运行）"""
import os
import shutil
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import kaoping_core as kc  # noqa: E402


def test_build_filename():
    # 默认模板
    assert kc.build_filename(kc.DEFAULT_NAME_PATTERN, 2026, 8, "张三") == \
        "作业长安全生产责任制履职履责考评表(8月张三).doc"
    # 全角/自定义模板
    assert kc.build_filename("{XXX}（{Y}年{X}月）", 2026, 12, "李四") == \
        "李四（2026年12月）.doc"
    # 非法字符替换
    assert kc.build_filename("{XXX}", 2026, 8, '王五/李六*') == "王五_李六_.doc"
    # Windows 保留设备名
    assert kc.build_filename("CON", 2026, 8, "x").startswith("_")
    assert kc.build_filename("NUL.{XXX}", 2026, 8, "x").startswith("_")
    # 结尾点/空格消毒
    assert kc.build_filename("{XXX}.doc", 2026, 8, "赵六. ") == "赵六.doc"


def test_replace_slot():
    body = "考评对象（作业长）：  孙忠      管理者姓名：x评价月份：                 评价人员签字"
    b = kc._replace_slot(body, "考评对象（作业长）：", "管理者姓名", "测试员")
    assert "：  测试员      管理者姓名" in b
    b = kc._replace_slot(b, "评价月份：", "评价人员签字", "8月")
    seg = b[b.find("评价月份：") + 5:b.find("评价人员签字")]
    assert seg.strip() == "8月", f"月份区间异常: {seg!r}"
    # 值居中：前后都有空白
    assert seg.startswith(" ") and seg.endswith(" ")


def test_item_rows():
    """行映射：13 个考评项，每项一个主填写行。"""
    assert kc.N_ITEMS == 13
    assert kc._item_rows(1) == [3]
    assert kc._item_rows(2) == [5]
    assert kc._item_rows(9) == [17]
    assert kc._item_rows(13) == [24]
    assert kc.TOTAL_ROW == 30


def test_sum_scores():
    items = {
        1: {"score": "20", "super_score": "20"},
        2: {"score": "5", "super_score": ""},
        3: {"score": "abc", "super_score": "5"},
    }
    # 第 1 项 20、第 2 项 5、第 3 项 abc 忽略 → 25
    assert kc._sum_scores(items, "score") == "25"
    # 上级：20+5 = 25
    assert kc._sum_scores(items, "super_score") == "25"
    assert kc._sum_scores({}, "score") == ""


def test_is_image():
    assert kc.is_image("a.JPG")
    assert kc.is_image("a.png")
    assert not kc.is_image("a.docx")
    assert not kc.is_image("a.xlsx")


def test_match_materials_file():
    assert kc.match_materials_file("扫雷/2026年扫雷统计表.xlsx") == 2
    assert kc.match_materials_file("综合大检查/安全综合检查通报.doc") == 3
    assert kc.match_materials_file("（炼铁）8月主题工作/1.安全专题会纪要.doc") == 13
    assert kc.match_materials_file("应急演练/皮带机演练方案.doc") == 8
    assert kc.match_materials_file("危险源/危险源控制点自查表.xlsx") == 6
    assert kc.match_materials_file("安全培训课件.pdf") == 4
    assert kc.match_materials_file("违章查处记录.xls") == 9
    assert kc.match_materials_file("危化品/一品一表.xlsx") == 7
    # 无匹配
    assert kc.match_materials_file("高温天气预警/8.3.jpg") is None
    assert kc.match_materials_file("领导视察/接待安排.docx") is None


def test_scan_materials_folder():
    tmp = tempfile.mkdtemp()
    try:
        os.makedirs(os.path.join(tmp, "扫雷"))
        os.makedirs(os.path.join(tmp, "应急演练"))
        # 匹配文件
        open(os.path.join(tmp, "扫雷", "扫雷统计表.xlsx"), "w").close()
        open(os.path.join(tmp, "应急演练", "演练方案.doc"), "w").close()
        # 不匹配文件
        open(os.path.join(tmp, "高温预警.jpg"), "w").close()
        # 应被跳过的考评表文件
        open(os.path.join(tmp, "作业长安全生产责任制履职履责考评表（张三8月）.doc"), "w").close()
        # 应被跳过的临时文件
        open(os.path.join(tmp, "~$临时.docx"), "w").close()

        result, unmatched = kc.scan_materials_folder(tmp)
        assert len(result[2]) == 1 and result[2][0].endswith("扫雷统计表.xlsx")
        assert len(result[8]) == 1 and result[8][0].endswith("演练方案.doc")
        assert len(result[1]) == 0
        assert len(unmatched) == 1, f"未匹配应为1个，实际 {len(unmatched)}"
        assert unmatched[0].endswith("高温预警.jpg")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_fit_size():
    # 已在 2cm 内（含浮点尾差）原样返回——绝不能进入缩放分支，
    # 否则再次 ScaleWidth=100 会把已缩好的对象放大回原始大小（曾实测 486×864pt）。
    assert kc._fit_size(31.9, 56.700001, 2.0) == (31.9, 56.700001)
    assert kc._fit_size(56.7, 56.7, 2.0) == (56.7, 56.7)
    assert kc._fit_size(56.4, 42.0, 2.0) == (56.4, 42.0)
    # 超限等比缩小，长边 ≤ 2cm（56.7pt）
    tw, th = kc._fit_size(486, 864, 2.0)
    assert tw <= 56.7 + 1e-6 and th <= 56.7 + 1e-6
    assert abs(tw / th - 486 / 864) < 0.01
    tw, th = kc._fit_size(2400, 1800, 2.0)
    assert tw <= 56.7 + 1e-6 and th <= 56.7 + 1e-6
    # 异常输入不抛错
    assert kc._fit_size(0, 100, 2.0) == (0, 100)
    assert kc._fit_size(-5, 100, 2.0) == (-5, 100)


def test_prepare_image_file():
    from PIL import Image
    tmp = tempfile.mkdtemp()
    try:
        big = os.path.join(tmp, "big.jpg")
        Image.new("RGB", (3000, 2000)).save(big)
        small = os.path.join(tmp, "small.png")
        Image.new("RGB", (800, 600)).save(small)
        p1, is_tmp1 = kc._prepare_image_file(big, tmp)
        assert is_tmp1 is True
        with Image.open(p1) as im:
            assert max(im.size) <= kc.MAX_EMBED_PX
        p2, is_tmp2 = kc._prepare_image_file(small, tmp)
        assert is_tmp2 is False and p2 == small
        # 非法图片回退原文件，不抛错
        bad = os.path.join(tmp, "bad.jpg")
        with open(bad, "w", encoding="utf-8") as f:
            f.write("not an image")
        p3, is_tmp3 = kc._prepare_image_file(bad, tmp)
        assert p3 == bad and is_tmp3 is False
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _make_fake_lvzhilv():
    """构造一个与真实月度履职履责表同构的临时 xlsx（班组长月度履职评价表）。"""
    import openpyxl
    tmp = tempfile.mkdtemp()
    p = os.path.join(tmp, "履职履责.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "作业长月度履职评价表"
    ws.cell(1, 1, "作业长（事业部）月度安全生产履职评价表")
    ws.cell(2, 1, "序号"); ws.cell(2, 2, "单位"); ws.cell(2, 3, "岗位"); ws.cell(2, 4, "姓名")
    ws.cell(2, 5, "考核维度（100分）"); ws.cell(2, 30, "得分\n排序")
    # 第 1 项 20 分、第 9 项 20 分、第 13 项 10 分，其余各 5 分（列 5,7,...,29）
    names = ["安全绩效", "隐患排查", "安全检查", "安全教育", "协力安全", "危险源管理",
             "危化品", "应急管理", "违章查处", "工机具", "劳防用品", "履职评价",
             "“六个一”工作实绩"]
    scores = [20, 5, 5, 5, 5, 5, 5, 5, 20, 5, 5, 5, 10]
    for j, (col, s) in enumerate(zip(range(5, 32, 2), scores)):
        ws.cell(3, col, "%d分" % s)
        ws.cell(4, col, names[j])
        ws.cell(4, col + 1, "扣分说明")
    ws.cell(5, 1, 1); ws.cell(5, 2, "兴达公司"); ws.cell(5, 3, "事业部作业长"); ws.cell(5, 4, "廖辉")
    for col, val in zip(range(5, 32, 2), [20, 5, 5, 5, 5, 5, 5, 5, 15, 5, 5, 5, 8]):
        ws.cell(5, col, val)
    ws.cell(5, 22, "未按违章查处")
    ws.cell(5, 30, 93)
    ws.cell(6, 4, "王旺")
    ws.cell(6, 5, 20); ws.cell(6, 7, 5)
    wb.save(p)
    return tmp, p


def test_read_eval_scores():
    tmp, p = _make_fake_lvzhilv()
    try:
        d = kc.read_eval_scores(p, "廖辉")
        assert d["name"] == "廖辉"
        assert d["total"] == "93"
        assert d["items"][1]["score"] == "20"
        assert d["items"][3]["score"] == "5"
        assert d["items"][9]["desc"] == ""
        assert d["items"][2]["desc"] == ""
        assert len(d["items"]) == 13
        # 未找到姓名：报错并列出可用姓名
        try:
            kc.read_eval_scores(p, "张三")
            assert False, "应抛出 ValueError"
        except ValueError as e:
            assert "廖辉" in str(e) and "王旺" in str(e)
        # 无目标 sheet：报错
        import openpyxl
        bad = os.path.join(tmp, "其他.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = "其他表"
        wb.save(bad)
        try:
            kc.read_eval_scores(bad, "廖辉")
            assert False, "应抛出 ValueError"
        except ValueError as e:
            assert "sheet" in str(e)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_list_eval_names():
    tmp, p = _make_fake_lvzhilv()
    try:
        assert kc.list_eval_names(p) == ["廖辉", "王旺"]
        # 无目标 sheet 时同样报错
        import openpyxl
        bad = os.path.join(tmp, "其他.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = "其他表"
        wb.save(bad)
        try:
            kc.list_eval_names(bad)
            assert False, "应抛出 ValueError"
        except ValueError as e:
            assert "sheet" in str(e)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_month_from_filename():
    assert kc._month_from_filename("安全员安全生产责任制履职清单考评表(孙忠8月).doc") == "8"
    assert kc._month_from_filename("作业长安全生产责任制履职履责考评表（廖辉7月）.doc") == "7"
    assert kc._month_from_filename("模板.doc") == ""


def test_excel_ext():
    assert kc._excel_ext(b"PK\x03\x04xxxx") == ".xlsx"        # xlsx 内容
    assert kc._excel_ext(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1") == ".xls"   # 旧版 xls 内容
    assert kc._excel_ext(b"") == ".xlsx"


def test_parse_kaoping_header():
    hdr = "考评对象（安全员）：孙忠  管理者姓名：（手签）评价月份：7月评价人员签字：（手签）\r\x07"
    name, month = kc._parse_kaoping_header(hdr)
    assert name == "孙忠" and month == "7"
    # 半角冒号 / 缺月份
    name2, month2 = kc._parse_kaoping_header("考评对象（安全员）: 王旺 管理者姓名:")
    assert name2 == "王旺" and month2 == ""
    assert kc._parse_kaoping_header("无关文本") == ("", "")


def test_extract_ole10native():
    # 构造 Ole10Native：4字节长度 + 0200 + GBK文件名 + NUL + 内容长度 + ZIP 内容
    content = b"PK\x03\x04" + b"hello" * 10 + b"PK\x05\x06" + b"\x00\x00"
    body = (b"\x02\x00" + "材料表.xlsx".encode("gbk") + b"\x00"
            + len(content).to_bytes(4, "little") + content)
    raw = len(body).to_bytes(4, "little") + body
    fn, ct = kc._extract_ole10native(raw)
    assert fn == "材料表.xlsx" and ct == content
    # OLE2(.doc) 内容
    ole_content = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 32
    body2 = (b"\x02\x00" + b"a.doc\x00"
             + len(ole_content).to_bytes(4, "little") + ole_content)
    raw2 = len(body2).to_bytes(4, "little") + body2
    fn2, ct2 = kc._extract_ole10native(raw2)
    assert fn2 == "a.doc" and ct2 == ole_content
    # 非法流
    assert kc._extract_ole10native(b"\x00" * 10) == (None, None)


def test_safe_filename():
    assert kc._safe_filename('a/b:c*d?e') == "a_b_c_d_e"
    assert kc._safe_filename("   ") == "未命名"


def test_sniff_office_kind():
    # xlsx：zip 内含 xl/
    import io
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", "")
        z.writestr("xl/workbook.xml", "")
    assert kc._sniff_office_kind(buf.getvalue()) == "excel"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", "")
        z.writestr("word/document.xml", "")
    assert kc._sniff_office_kind(buf.getvalue()) == "word_docx"
    assert kc._sniff_office_kind(b"D0CF11E0A1B11AE1") == "file"
    assert kc._sniff_office_kind(b"") == "file"
    assert kc._sniff_office_kind(None) == "file"


def test_ole_kind_match():
    assert kc._ole_kind_match("excel", "excel") is True
    assert kc._ole_kind_match("excel", "file") is True
    assert kc._ole_kind_match("excel", "pkg") is False
    assert kc._ole_kind_match("word_docx", "pkg") is True
    assert kc._ole_kind_match("pkg", "excel") is True


def test_self_check():
    """岗位参数自检应通过（行映射/总分行/权重键自洽）。"""
    kc.self_check()


def _first_item_row(idx):
    """取考评项主行（兼容单行 int 与双行 list 两种 ITEM_ROWS 值）。"""
    v = kc.ITEM_ROWS[idx]
    return v[0] if isinstance(v, (list, tuple)) else v


class _FakeRange:
    """极简 Range 替身：Text 读写委托给所属单元格；Collapse/End 仅为接口占位。"""

    def __init__(self, owner):
        self._owner = owner
        self.End = 1
        self.InlineShapes = owner.inline_shapes

    @property
    def Text(self):
        return self._owner.text

    @Text.setter
    def Text(self, value):
        self._owner.text = value

    def Collapse(self, direction=0):
        return None

    def InsertParagraphAfter(self):
        return None


class _FakeCell:
    def __init__(self, text=""):
        self.text = text
        self.inline_shapes = []
        self.Range = _FakeRange(self)


class _FakeTable:
    """假 Word 表格：Cell(r, c) 返回可写的假单元格。"""

    def __init__(self, rows, cols):
        self.Rows = type("Rows", (), {"Count": rows})()
        self.Columns = type("Cols", (), {"Count": cols})()
        self._cells = {}

    def Cell(self, row, col):
        key = (row, col)
        if key not in self._cells:
            self._cells[key] = _FakeCell("")
        return self._cells[key]


class _FakeTables:
    """假 Tables 集合：doc.Tables(1) 可调用返回指定表格。"""

    def __init__(self, table):
        self._table = table

    def __call__(self, index=1):
        return self._table


class _FakeDoc:
    def __init__(self, table):
        self.Tables = _FakeTables(table)
        self.saved = None

    def SaveAs2(self, path, fmt=0):
        self.saved = (path, fmt)

    def SaveAs(self, path, fmt=0):
        self.saved = (path, fmt)

    def Close(self, save=False):
        pass


class _FakeWord:
    """假 Word.Application：仅支撑 generate_doc 用到的接口（不需要真实 Word）。"""

    def __init__(self, doc):
        self.doc = doc
        self.Visible = True
        self.DisplayAlerts = -1

    @property
    def Documents(self):
        return self

    def Open(self, path):
        return self.doc

    def Quit(self):
        pass


def test_generate_doc_fill_sequence():
    """Mock Word COM：验证生成流程填充顺序（表头 -> 逐项 -> 合计 -> 保存），无需本机 Word。

    覆盖：模板结构校验入口、表头姓名/月份占位替换、逐项内容写入、合计写入、SaveAs2 输出。
    """
    tmp = tempfile.mkdtemp()
    try:
        tb = _FakeTable(kc.TOTAL_ROW, 10)
        header_cell = tb.Cell(kc.HEADER_ROW, 1)
        if getattr(kc, "HEADER_ROLE_NEXT", "") == "评价月份":
            # 主要负责人模板表头无「管理者姓名」栏
            header_cell.text = (kc.HEADER_ROLE_LABEL + "  孙忠   "
                                + kc.HEADER_MONTH_LABEL + "                 评价人员签字：（手签）")
        else:
            header_cell.text = (kc.HEADER_ROLE_LABEL + "  孙忠      管理者姓名：（手签）"
                                + kc.HEADER_MONTH_LABEL + "                 评价人员签字：（手签）")
        fake_doc = _FakeDoc(tb)
        orig_new_word = kc._new_word_app
        kc._new_word_app = lambda: _FakeWord(fake_doc)
        try:
            items = {}
            for idx in range(1, kc.N_ITEMS + 1):
                items[idx] = {"desc": "d%d" % idx, "score": "5", "material_text": "",
                              "materials": [], "eval_desc": "e%d" % idx, "super_score": "4"}
            items[1]["score"] = "10"
            items[1]["super_score"] = "9"
            out = os.path.join(tmp, "out.doc")
            kc.generate_doc("模板.doc", out, "张三", 8, items, year=2026)
        finally:
            kc._new_word_app = orig_new_word
        # 表头：姓名与月份已替换
        assert "张三" in header_cell.text and "8月" in header_cell.text
        # 逐项内容写入对应列（第 1 项主行）
        r1 = _first_item_row(1)
        assert tb.Cell(r1, kc.COL_SELF_DESC).text == "d1"
        assert tb.Cell(r1, kc.COL_SELF_SCORE).text == "10"
        assert tb.Cell(r1, kc.COL_EVAL_DESC).text == "e1"
        assert tb.Cell(r1, kc.COL_EVAL_SCORE).text == "9"
        # 合计写入 TOTAL_ROW 的自评/上级列
        self_total = 10 + 5 * (kc.N_ITEMS - 1)
        super_total = 9 + 4 * (kc.N_ITEMS - 1)
        assert tb.Cell(kc.TOTAL_ROW, kc.COL_SELF_DESC).text == str(self_total)
        assert tb.Cell(kc.TOTAL_ROW, kc.COL_EVAL_DESC).text == str(super_total)
        # 已按输出路径调用 SaveAs2
        assert fake_doc.saved == (out, 0)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)



def main():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failed = 0
    for t in tests:
        try:
            t()
            print(f"[PASS] {t.__name__}")
        except AssertionError as e:
            failed += 1
            print(f"[FAIL] {t.__name__}: {e}")
        except Exception as e:
            failed += 1
            print(f"[ERROR] {t.__name__}: {type(e).__name__}: {e}")
    print(f"\n共 {len(tests)} 项测试，失败 {failed} 项")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
