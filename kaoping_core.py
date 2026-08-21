#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""安全员履职考评表生成核心逻辑 v1.1.2（兼容 Windows 7 SP1 ~ Windows 11）"""

import json
import logging
import os
import re
import shutil
import sys
import tempfile
import time
from datetime import datetime

import win32com.client
from win32com.client import gencache

TEMPLATE_NAME = "安全员安全生产责任制履职清单考评表（模板）.doc"


ITEM_ROWS = {
    1: 3, 2: 5, 3: 6, 4: 7, 5: 8, 6: 9,
    7: 10, 8: 11, 9: 12, 10: 13, 11: 14, 12: 15,
}
TOTAL_ROW = 16
HEADER_ROW = 1
COL_SELF_DESC = 6
COL_SELF_SCORE = 7
COL_MATERIAL = 8
COL_EVAL_DESC = 9
COL_EVAL_SCORE = 10
IMG_MAX_SIZE_CM = 2.0      # 图片/OLE 对象显示长宽上限（等比缩放，长宽均不超过该值）
DEFAULT_NAME_PATTERN = "安全员安全生产责任制履职清单考评表({X}月{XXX}).doc"


def find_template(template_path=None):
    """定位模板文件：显式路径 > exe/脚本目录 > 上一级目录 > _MEIPASS(旧版兼容)。

    模板为外置文件：打包成 exe 后需与 exe 同目录（或上一级目录）分发，
    不再内置到 exe 中；源码运行时按脚本所在目录定位。
    """
    if template_path and os.path.exists(template_path):
        return template_path

    # 打包(exe)时以 exe 所在目录为基准；源码运行时以脚本所在目录为基准
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))

    candidates = [
        os.path.join(base, TEMPLATE_NAME),
        os.path.join(os.path.dirname(base), TEMPLATE_NAME),
    ]
    # 兼容旧版打包：模板仍被内置在 exe 资源目录(_MEIPASS)中的情况
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(os.path.join(meipass, TEMPLATE_NAME))
    for c in candidates:
        if c and os.path.exists(c):
            return c
    raise FileNotFoundError(
        f"未找到模板文件 '{TEMPLATE_NAME}'，请确保模板在以下目录之一：{chr(10)}"
        f"  - {base}{chr(10)}"
        f"  - {os.path.dirname(base)}"
    )


def build_filename(pattern, year, month, name, default_ext=".doc"):
    """按命名模板生成文件名。占位符：{Y}=年, {X}=月, {XXX}=姓名"""
    fn = pattern.strip()
    if not fn:
        fn = DEFAULT_NAME_PATTERN
    fn = fn.replace("{Y}", str(year)).replace("{年}", str(year))
    fn = fn.replace("{X}", str(month)).replace("{月份}", str(month))
    fn = fn.replace("{XXX}", str(name)).replace("{姓名}", str(name))
    fn = re.sub(r'[\\/:*?"<>|]', "_", fn)
    # 分离扩展名，只对主干做 Windows 消毒（去掉结尾点/空格）
    if fn.lower().endswith(".docx"):
        stem, ext = fn[:-5], fn[-5:]
    elif fn.lower().endswith(".doc"):
        stem, ext = fn[:-4], fn[-4:]
    else:
        stem, ext = fn, default_ext
    stem = stem.rstrip(" .") or "未命名"
    fn = stem + ext
    # Windows 保留设备名（CON/PRN/AUX/NUL/COM1~9/LPT1~9），按第一个点前的主名判断
    stem0 = fn.split(".", 1)[0].upper()
    reserved = {"CON", "PRN", "AUX", "NUL"} | {f"COM{i}" for i in range(1, 10)} \
        | {f"LPT{i}" for i in range(1, 10)}
    if stem0 in reserved:
        fn = "_" + fn
    return fn


# ============ 支撑材料文件夹自动匹配 ============
# 每个考评项的关键词表（按文件名/相对路径中的子文件夹名匹配）。
# 第一词为该考评项标题词（权重 2），其余为扩展词（权重 1）；
# 界面“编辑权重”按钮可打开本地 material_rules.json 调整（修改后重启生效）。
# 关键词基于各月实际文件夹命名归纳（如 扫雷→⑩、安全月报→⑫、安监网/安检网录入→⑪）。
ITEM_MATCH_RULES = {
    1: ["安全绩效", "绩效", "考核", "事故", "工伤", "轻伤", "月度会"],
    2: ["管理体系", "体系", "制度", "责任制", "标准化"],
    3: ["教育培训", "培训", "教育", "学习", "课件", "交底", "考试", "双考", "应知", "安全活动"],
    4: ["会议活动", "会议", "班会", "班前会", "活动", "纪要", "部署", "发言"],
    5: ["检查改进", "检查", "巡查", "督查", "整改", "找茬", "大检查", "自查", "排查表"],
    6: ["应急演习", "应急", "演练", "演习", "预案"],
    7: ["监督执行", "监督", "值守", "带班", "值班", "旁站", "应急值守"],
    8: ["安全能力", "能力", "取证", "特种作业", "资格证", "上岗证", "特种工", "证书", "规程", "复审"],
    9: ["危险源", "风险", "危害", "辨识", "评估"],
    10: ["隐患排查", "隐患", "排查", "扫雷", "有限空间"],
    11: ["违章管理", "违章", "违规", "三违", "违纪", "反思", "安监网", "安检网", "禁令"],
    12: ["其它", "其他", "主题", "总结", "小结", "月报", "台账", "人数", "体检", "协议", "资料",
         "汇报", "保险", "方案", "通知", "清单"],
}
# 扫描时跳过：临时文件、考评表/履责表自身、无意义扩展名
SKIP_FILE_PARTS = ("~$",)
SKIP_NAME_PARTS = ("考评表", "履职履责表", "履职评价表", "履责表", "履职履责")
SKIP_EXTS = {".lnk", ".url", ".ini", ".tmp", ".db", ".py", ".pyc", ".md"}

# 本地文件名权重配置文件（与 exe/脚本同目录，首次运行自动生成）
MATERIAL_RULES_FILE = "material_rules.json"


def ensure_material_rules_file(base_dir=None):
    """确保本地文件名权重配置文件存在（首次自动生成），返回其路径。"""
    base_dir = base_dir or os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base_dir, MATERIAL_RULES_FILE)
    if not os.path.exists(path):
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({str(k): list(v) for k, v in ITEM_MATCH_RULES.items()},
                          f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    return path


def load_material_rules(path=None):
    """读取本地文件名权重配置；文件缺失/非法时回退内置 ITEM_MATCH_RULES。"""
    if path and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            rules = {}
            for k, v in data.items():
                idx = int(str(k).split(".")[0])          # 兼容 "1" / "1. 安全绩效"
                if isinstance(v, dict):                  # 兼容 {"weight":2,"keywords":[...]}
                    kw = v.get("keywords") or v.get("关键词") or []
                else:
                    kw = v
                kw = [str(x).strip() for x in (kw or []) if str(x).strip()]
                if kw:
                    rules[idx] = kw
            if rules:
                return rules
        except Exception:
            pass
    return dict(ITEM_MATCH_RULES)


def score_materials_file(rel_path, rules):
    """按权重打分并返回 (考评项编号, 得分)；无匹配返回 (None, 0)。
    每个考评项第 1 词权重 2、其余词权重 1；命中分高者优先，同分取编号小者。"""
    text = rel_path.rsplit(".", 1)[0]                    # 去扩展名，保留相对路径（含子文件夹名）
    scores = {}
    for idx, kws in rules.items():
        s = 0
        for i, kw in enumerate(kws):
            if kw and kw in text:
                s += 2 if i == 0 else 1
        if s:
            scores[idx] = s
    if not scores:
        return None, 0
    idx = max(scores, key=lambda k: (scores[k], -k))
    return idx, scores[idx]


def match_materials_file(rel_path, rules=None):
    """按关键词把文件相对路径匹配到考评项，返回编号(1~12)；无匹配返回 None。"""
    idx, _score = score_materials_file(rel_path, rules or load_material_rules())
    return idx


def find_month_subfolder(folder, year, month):
    """在年度/根目录中自动定位当月子文件夹（如 2026.08 / 2026.8 / 8月）。
    找到返回其路径，否则原样返回 folder。"""
    targets = {f"{year}.{month:02d}", f"{year}.{month}", f"{year}-{month:02d}",
               f"{year}-{month}", f"{year}年{month}月", f"{month}月"}
    try:
        for name in os.listdir(folder):
            full = os.path.join(folder, name)
            if os.path.isdir(full) and name in targets:
                return full
    except Exception:
        pass
    return folder


def scan_materials_folder(folder, max_per_item=15, rules=None):
    """递归扫描支撑材料文件夹，按文件名权重自动匹配到各考评项。

    返回 (result, unmatched)：
      result   = {1..12: [文件绝对路径...]}（每项最多 max_per_item 个，按权重分降序）
      unmatched= [未匹配(或超限)的文件绝对路径...]
    """
    folder = os.path.abspath(folder)
    if not os.path.isdir(folder):
        raise ValueError("支撑材料文件夹不存在：" + folder)
    rules = rules or load_material_rules()
    buckets = {i: [] for i in range(1, 13)}              # {idx: [(score, path)...]}
    unmatched = []
    for root, _dirs, files in os.walk(folder):
        for fn in sorted(files):
            if fn.startswith(SKIP_FILE_PARTS):
                continue
            if any(p in fn for p in SKIP_NAME_PARTS):
                continue
            if os.path.splitext(fn)[1].lower() in SKIP_EXTS:
                continue
            full = os.path.join(root, fn)
            rel = os.path.relpath(full, folder)
            idx, score = score_materials_file(rel, rules)
            if idx is None:
                unmatched.append(full)
            elif len(buckets[idx]) < max_per_item:
                buckets[idx].append((score, full))
            else:
                unmatched.append(full)  # 超过每项上限，按未匹配提示用户手动处理
    result = {i: [p for _s, p in sorted(bucket, key=lambda t: -t[0])]
              for i, bucket in buckets.items()}
    return result, unmatched


# ============ 月度履职履责表(xlsx) 评分自动读取 ============
# 目标 sheet：「安全员月度履职评价表」（关键字匹配，排除「评价标准」）。
# 布局（2026 年 6/7/8 月多份实测，兴达/铁厂版一致）：
#   表头行为含「扣分说明」的行；偶数列为 12 个考评项名、奇数列为其扣分说明；
#   姓名在「姓名」列（D 列）；得分合计在含「得分」「合计」的表头列（AC 列）。
XLSX_SHEET_INCLUDE = ("安全员", "评价")
XLSX_SHEET_EXCLUDE = ("标准",)
# 12 个考评项在 xlsx 表头中的匹配关键词（任一命中即可）
XLSX_ITEM_KEYWORDS = {
    1: ("安全绩效",),
    2: ("管理体系", "体系"),
    3: ("教育培训", "培训"),
    4: ("会议活动", "会议"),
    5: ("检查改进", "检查"),
    6: ("应急演习", "应急", "演习", "演练"),
    7: ("监督执行", "监督"),
    8: ("安全能力", "能力"),
    9: ("危险源", "危险源", "风险"),
    10: ("隐患排查", "隐患"),
    11: ("违章管理", "违章"),
    12: ("其它", "其他"),
}


def read_eval_scores(xlsx_path, name):
    """从月度履职履责表 xlsx 读取指定姓名安全员的 12 项评分与扣分说明。

    返回 {"name": 姓名, "items": {1..12: {"score": str, "desc": str}},
           "total": 总分(str 或 "")}。
    未找到 sheet / 表头 / 姓名时抛出 ValueError 并附可用信息。
    """
    import openpyxl

    def _norm(v):
        return "".join(str(v).split()) if v is not None else ""

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = None
    for sn in wb.sheetnames:
        if (all(k in sn for k in XLSX_SHEET_INCLUDE)
                and not any(k in sn for k in XLSX_SHEET_EXCLUDE)):
            sheet = wb[sn]
            break
    if sheet is None:
        raise ValueError("未找到「安全员月度履职评价表」sheet，可用：%s"
                         % "、".join(wb.sheetnames))

    # 1) 表头行：含「扣分说明」的行
    header_row = None
    for r in range(1, min(sheet.max_row, 8) + 1):
        if any("扣分说明" in _norm(sheet.cell(row=r, column=c).value)
               for c in range(1, sheet.max_column + 1)):
            header_row = r
            break
    if header_row is None:
        raise ValueError("未找到考评项表头行（含「扣分说明」），文件格式可能已变化")

    # 2) 姓名列 / 得分合计列（在表头行之前找标签）
    name_col, total_col = None, None
    for r in range(1, header_row):
        for c in range(1, sheet.max_column + 1):
            h = _norm(sheet.cell(row=r, column=c).value)
            if h == "姓名" and name_col is None:
                name_col = c
            if "得分" in h and total_col is None:
                total_col = c
    if name_col is None:
        name_col = 4  # 兜底：D 列

    # 3) 项名列：表头按关键词匹配 12 个考评项
    header_vals = {c: _norm(sheet.cell(row=header_row, column=c).value)
                   for c in range(1, sheet.max_column + 1)}
    item_cols = {}
    for idx in range(1, 13):
        kws = XLSX_ITEM_KEYWORDS[idx]
        hit = [c for c, h in header_vals.items()
               if h and any(k in h for k in kws) and c not in item_cols.values()]
        if not hit:
            raise ValueError("表头中未找到第 %d 项「%s」，文件格式可能已变化"
                             % (idx, ITEM_MATCH_RULES[idx][0]))
        item_cols[idx] = hit[0]

    # 4) 定位姓名行
    person_row, names = None, []
    for r in range(header_row + 1, sheet.max_row + 1):
        nm = sheet.cell(row=r, column=name_col).value
        if nm is None or not str(nm).strip():
            continue
        names.append(str(nm).strip())
        if str(nm).strip() == name.strip():
            person_row = r
            break
    if person_row is None:
        raise ValueError("文件中未找到姓名「%s」，可用姓名：%s"
                         % (name, "、".join(names) if names else "(空)"))

    # 5) 读取 12 项评分与扣分说明、总分
    items = {}
    for idx, col in item_cols.items():
        items[idx] = {
            "score": _norm(sheet.cell(row=person_row, column=col).value),
            "desc": _norm(sheet.cell(row=person_row, column=col + 1).value),
        }
    total = _norm(sheet.cell(row=person_row, column=total_col).value) if total_col else ""
    return {"name": name.strip(), "items": items, "total": total}


# ============ 已生成考评表(.doc) 读取（评分/评价/支撑材料） ============
_ZIP_MAGIC = b"PK\x03\x04"
_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _safe_filename(name):
    """Windows 文件名消毒。"""
    name = re.sub(r'[\\/:*?"<>|]', "_", str(name)).strip(" .")
    return name or "未命名"


def _parse_kaoping_header(header_text):
    """解析考评表表头 R1，返回 (姓名, 月份数字串)。"""
    body = (header_text or "").replace("\x07", "").rstrip("\r")

    def _slot(text, label, next_label):
        li = text.find(label)
        if li < 0:
            return ""
        start = li + len(label)
        for ch in ("：", ":"):
            ci = text.find(ch, start)
            if ci >= 0:
                start = ci + 1
                break
        ni = text.find(next_label, start)
        return "".join(text[start:ni if ni > 0 else len(text)].split())

    name = _slot(body, "考评对象", "管理者姓名")
    month = _slot(body, "评价月份", "评价人员").rstrip("月").strip()
    return name, month


def _cell_text(tb, row, col):
    """读取表格单元格纯文本。"""
    try:
        return tb.Cell(row, col).Range.Text.replace("\x07", "").replace("\r", "").strip()
    except Exception:
        return ""


def _export_inline_picture(shp, out_path):
    """将内嵌图片保存为 PNG（CopyAsPicture -> 剪贴板 -> Pillow）。返回是否成功。"""
    from PIL import Image, ImageGrab
    try:
        shp.Range.CopyAsPicture()
        img = ImageGrab.grabclipboard()
        if isinstance(img, Image.Image):
            img.save(out_path)
            return os.path.exists(out_path)
    except Exception:
        pass
    return False


def _export_ole_via_com(shp, idx, no, out_dir, iconlabel=""):
    """尝试用 Word COM 直接导出原生 OLE 对象（Word/Excel 可 SaveAs）。

    返回 (kind, 已保存路径或 None)。kind 供 ObjectPool 分组对应：
    'excel' / 'word_doc' / 'word_docx' / 'pkg'。
    """
    try:
        progid = shp.OLEFormat.ProgID or ""
    except Exception:
        progid = ""
    if "Excel" in progid:
        kind = "excel"
    elif "Word.Document.12" in progid:
        kind = "word_docx"
    elif "Word.Document" in progid:
        kind = "word_doc"
    else:
        return "pkg", None
    try:
        obj = shp.OLEFormat.Object
        base = _safe_filename(iconlabel) if iconlabel and iconlabel.strip() else "项%d_材料%d" % (idx, no)
        if kind == "excel":
            out = os.path.join(out_dir, os.path.splitext(base)[0] + ".xlsx")
            obj.SaveAs(out, 51)
            return kind, out
        ext = "docx" if kind == "word_docx" else "doc"
        out = os.path.join(out_dir, os.path.splitext(base)[0] + "." + ext)
        save = getattr(obj, "SaveAs2", obj.SaveAs)
        save(out, 12 if kind == "word_docx" else 0)
        return kind, out
    except Exception:
        return kind, None


def _ole_kind_match(want, have):
    """ObjectPool 条目 kind 是否满足 OLE 形状的 kind 需求。"""
    if want == "excel":
        return have in ("excel", "file")
    if want in ("word_docx", "word_doc"):
        return have in ("word_docx", "word_native", "pkg", "file")
    # 'pkg'（通用包）可匹配任何可提取类型
    return have in ("pkg", "file", "excel", "word_docx")


def _sniff_office_kind(content):
    """按内容识别 package 流是 xlsx 还是 docx（zip 内 xl/ 或 word/ 目录）。"""
    if not content or content[:2] != b"PK":
        return "file"
    try:
        import io
        import zipfile
        names = zipfile.ZipFile(io.BytesIO(content)).namelist()
    except Exception:
        return "file"
    if any(n.startswith("xl/") for n in names):
        return "excel"
    if any(n.startswith("word/") for n in names):
        return "word_docx"
    return "file"


def _extract_ole10native(raw):
    """从 \x01Ole10Native 流提取原始文件，返回 (原始文件名, 文件内容 bytes) 或 (None, None)。"""
    off = 0
    if len(raw) > 4 and int.from_bytes(raw[0:4], "little") == len(raw) - 4:
        off = 4
    if raw[off:off + 2] != b"\x02\x00":
        return None, None
    p = off + 2
    e = raw.find(b"\x00", p)
    fname = raw[p:e].decode("gbk", "ignore").strip() if e >= 0 else ""
    for magic in (_ZIP_MAGIC, _OLE_MAGIC):
        cstart = raw.find(magic)
        if cstart < 4:
            continue
        clen = int.from_bytes(raw[cstart - 4:cstart], "little")
        if 0 < clen <= len(raw) - cstart:
            return fname or None, raw[cstart:cstart + clen]
    return fname or None, None


def _read_doc_object_pool(doc_path):
    """用 olefile 读取 .doc 的 ObjectPool，返回按文档顺序的
    [(storage名, kind, 文件字节, 原始文件名)]。
    kind：'excel'(package流) / 'word_docx'(package流+Word) / 'pkg'(Ole10Native) /
          'word_native'(原生 WordDocument 流，需 COM 导出) / 'unknown'。
    """
    import re as _re

    import olefile
    ole = olefile.OleFileIO(doc_path)
    entries = []
    try:
        seen = set()
        for path in ole.listdir():
            if len(path) >= 2 and path[0] == "ObjectPool":
                name = path[1]
                if name in seen:
                    continue
                seen.add(name)
                progid = ""
                if ole.exists(["ObjectPool", name, "\x01CompObj"]):
                    data = ole.openstream(["ObjectPool", name, "\x01CompObj"]).read()
                    m = _re.search(rb"[\x20-\x7e]{6,}", data)
                    if m:
                        progid = m.group(0).decode("ascii", "ignore")
                content, fname, kind = None, "", "unknown"
                if ole.exists(["ObjectPool", name, "package"]):
                    content = ole.openstream(["ObjectPool", name, "package"]).read()
                    if "Excel" in progid:
                        kind = "excel"
                    elif "Word" in progid:
                        kind = "word_docx"
                    else:
                        kind = _sniff_office_kind(content)
                elif ole.exists(["ObjectPool", name, "\x01Ole10Native"]):
                    raw = ole.openstream(["ObjectPool", name, "\x01Ole10Native"]).read()
                    fname, content = _extract_ole10native(raw)
                    kind = "pkg"
                elif ole.exists(["ObjectPool", name, "WordDocument"]):
                    kind = "word_native"
                entries.append((name, kind, content, fname))
    finally:
        ole.close()
    return entries


def extract_kaoping_doc(doc_path, out_dir=None, progress_cb=None):
    """读取已生成的考评表 .doc：姓名/月份/12 项自评与上级评价/支撑材料。

    支撑材料（图片/OLE 内嵌文件）提取保存到 out_dir（默认 .doc 同目录"提取材料"）。
    返回 {"name", "month", "items": {1..12: {"desc","score","material_text",
           "materials":[文件路径], "eval_desc","super_score"}}, "warnings":[str]}。
    """
    doc_path = os.path.abspath(doc_path)
    out_dir = out_dir or os.path.join(os.path.dirname(doc_path), "提取材料")
    os.makedirs(out_dir, exist_ok=True)
    warnings = []

    # ---- 阶段1：Word COM 读取文字 + 提取图片 + 记录 OLE 出现顺序 ----
    word = _new_word_app()
    word.Visible = False
    word.DisplayAlerts = 0
    doc = None
    items = {}
    mat_records = []  # 已提取材料：(考评项 idx, 形状内序号, 文件路径)
    ole_shape_items = []  # 需阶段2提取的 OLE 对象：每元素 = (考评项 idx, kind, 形状内序号, iconlabel文件名)
    try:
        doc = word.Documents.Open(doc_path)
        tb = doc.Tables(1)
        if tb.Rows.Count != TOTAL_ROW or tb.Columns.Count != 10:
            raise ValueError("表格结构异常：期望 %d 行×10 列，实际 %d 行×%d 列"
                             % (TOTAL_ROW, tb.Rows.Count, tb.Columns.Count))
        header_text = tb.Cell(HEADER_ROW, 1).Range.Text
        name, month = _parse_kaoping_header(header_text)
        for idx in range(1, 13):
            row = ITEM_ROWS[idx]
            items[idx] = {
                "desc": _cell_text(tb, row, COL_SELF_DESC),
                "score": _cell_text(tb, row, COL_SELF_SCORE),
                "material_text": "",
                "materials": [],
                "eval_desc": _cell_text(tb, row, COL_EVAL_DESC),
                "super_score": _cell_text(tb, row, COL_EVAL_SCORE),
            }
            mat_cell = tb.Cell(row, COL_MATERIAL)
            mat_txt = (mat_cell.Range.Text.replace("\x07", "")
                       .replace("\r", "").replace("\x01", "").strip())
            items[idx]["material_text"] = mat_txt
            mat_no = 0
            for s in list(mat_cell.Range.InlineShapes):
                mat_no += 1
                if s.Type == 3:  # 图片
                    out = os.path.join(out_dir, "项%d_图%d.png" % (idx, mat_no))
                    if _export_inline_picture(s, out):
                        mat_records.append((idx, mat_no, out))
                    else:
                        warnings.append("第 %d 项第 %d 张图片提取失败" % (idx, mat_no))
                else:            # OLE 对象：原生 Office 尝试 COM 导出，其余交阶段2
                    try:
                        _label = s.OLEFormat.IconLabel or ""
                    except Exception:
                        _label = ""
                    _kind, saved = _export_ole_via_com(s, idx, mat_no, out_dir, _label)
                    if saved:
                        mat_records.append((idx, mat_no, saved))
                    else:
                        ole_shape_items.append((idx, _kind, mat_no, _label))
            if progress_cb:
                try:
                    progress_cb(idx)
                except Exception:
                    pass
    finally:
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        word.Quit()

    # ---- 阶段2：olefile 提取未由 COM 导出的 OLE 内嵌文件 ----
    try:
        entries = _read_doc_object_pool(doc_path)
    except Exception as e:
        warnings.append("OLE 材料解析失败：" + str(e))
        entries = []
    if entries and ole_shape_items:
        # 本工具生成的文档包含原生 Word OLE（word_native/word_docx），其 ObjectPool 中
        # Excel 条目顺序与文档形状顺序相反（多份实测一致）；原版手工文档均为
        # Package(Ole10Native)/Excel(package流)，顺序一致，按正序对应。
        has_native_ole = any(entries[p][1] in ("word_native", "word_docx")
                             for p in range(len(entries)))
        excel_shapes = [r for r in ole_shape_items if r[1] == "excel"]
        excel_entries = [p for p in range(len(entries))
                         if entries[p][1] == "excel" and entries[p][2]]
        if has_native_ole:
            excel_entries = list(reversed(excel_entries))
        used = set()
        if excel_shapes:
            for (idx, _kind, mat_no, label), epos in zip(excel_shapes, excel_entries):
                if epos in used:
                    continue
                used.add(epos)
                _ename, kind, content, fname = entries[epos]
                base = (os.path.splitext(_safe_filename(label))[0]
                        if label and label.strip() else "项%d_材料%d" % (idx, mat_no))
                out = os.path.join(out_dir, base + ".xlsx")
                try:
                    with open(out, "wb") as f:
                        f.write(content)
                    mat_records.append((idx, mat_no, out))
                except Exception as e:
                    warnings.append("第 %d 项材料保存失败：%s" % (idx, e))
        # 其余类型（Package 等）按顺序对应
        for idx, want_kind, mat_no, label in ole_shape_items:
            if want_kind == "excel":
                continue
            chosen = None
            for pos in range(len(entries)):
                if pos in used or not entries[pos][2]:
                    continue
                if _ole_kind_match(want_kind, entries[pos][1]):
                    chosen = pos
                    break
            if chosen is None:
                warnings.append("第 %d 项 OLE 材料未能对应到内嵌文件" % idx)
                continue
            used.add(chosen)
            _ename, kind, content, fname = entries[chosen]
            if label and label.strip():
                fname_out = _safe_filename(label)
            elif kind == "word_docx":
                fname_out = "项%d_材料%d.docx" % (idx, mat_no)
            else:
                base = os.path.basename(fname) or "项%d_材料%d" % (idx, mat_no)
                fname_out = base if os.path.splitext(base)[1] else base + ".bin"
            out = os.path.join(out_dir, _safe_filename(fname_out))
            try:
                with open(out, "wb") as f:
                    f.write(content)
                mat_records.append((idx, mat_no, out))
            except Exception as e:
                warnings.append("第 %d 项材料保存失败：%s" % (idx, e))

    # 按形状内序号归并材料（保持原文档中图片/文件的先后顺序）
    for idx in range(1, 13):
        items[idx]["materials"] = [p for _i, _n, p in
                                   sorted((r for r in mat_records if r[0] == idx),
                                          key=lambda r: r[1])]

    return {"name": name, "month": month, "items": items, "warnings": warnings}


def _set_cell_text(cell, text):
    """写入单元格文本，保留模板原有字体/段落格式（Range 替换法）。"""
    text = "" if text is None else str(text)
    rng = cell.Range
    rng.End = rng.End - 1
    rng.Text = text


def _clear_cell_content(cell):
    """清空单元格内文字与图片，保留段落结构。"""
    for shp in list(cell.Range.InlineShapes):
        shp.Delete()
    rng = cell.Range
    rng.End = rng.End - 1
    rng.Text = ""


def _cm2pt(v):
    return v * 28.35


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tif", ".tiff"}


def is_image(path):
    """判断文件是否为可内嵌图片。"""
    return os.path.splitext(path)[1].lower() in IMAGE_EXTS


MAX_EMBED_PX = 1600  # 图片嵌入文档前的最大边长（像素），压缩超大照片避免 .doc 体积失控
FIT_TOLERANCE_PT = 0.5  # 尺寸达标容差(pt)：Word 返回宽高常带浮点尾差（如 56.700001）


def _fit_size(w, h, max_cm):
    """计算在 max_cm(厘米) 限制内的目标尺寸(pt)。

    已达标（含浮点容差）原样返回。调用方以返回值是否变化来判断是否需要缩放；
    对象本身已在 2cm 内时绝不能进入缩放分支——否则再次调用
    ScaleWidth/ScaleHeight=100 会把已缩好的对象放回原始大小，破坏版式
    （曾实测：2cm 内图片被还原成 486×864pt，撑爆表格）。
    """
    max_pt = _cm2pt(max_cm)
    if w <= 0 or h <= 0:
        return w, h
    if w <= max_pt + FIT_TOLERANCE_PT and h <= max_pt + FIT_TOLERANCE_PT:
        return w, h
    scale = min(max_pt / w, max_pt / h)
    return round(w * scale, 1), round(h * scale, 1)


def _prepare_image_file(path, tmp_dir):
    """图片边长超过 MAX_EMBED_PX 时，等比压缩到 tmp_dir 下的临时文件。

    返回 (嵌入路径, 是否为临时文件)。压缩只影响文档体积，显示尺寸仍由
    IMG_MAX_SIZE_CM 控制；任何异常都回退原文件，不影响主流程。
    """
    from PIL import Image

    try:
        with Image.open(path) as im:
            w, h = im.size
        if w <= MAX_EMBED_PX and h <= MAX_EMBED_PX:
            return path, False
        scale = min(MAX_EMBED_PX / w, MAX_EMBED_PX / h)
        nw, nh = max(1, int(w * scale)), max(1, int(h * scale))
        is_png = os.path.splitext(path)[1].lower() == ".png"
        suffix = ".png" if is_png else ".jpg"
        fd, tmp = tempfile.mkstemp(suffix=suffix, dir=tmp_dir)
        os.close(fd)
        with Image.open(path) as im:
            im = im.resize((nw, nh), Image.LANCZOS)
            if is_png:
                im.save(tmp, "PNG")
            else:
                im.convert("RGB").save(tmp, "JPEG", quality=88)
        return tmp, True
    except Exception:
        return path, False


def _insert_image(cell, path, tmp_dir=None):
    """在单元格末尾插入单张图片（内嵌），自动等比缩放，长宽均不超过 IMG_MAX_SIZE_CM。

    tmp_dir 非空时先对超大图片做像素压缩（控制 .doc 体积），再插入。
    """
    from PIL import Image

    if tmp_dir:
        path, _tmp = _prepare_image_file(path, tmp_dir)
    with Image.open(path) as im:
        ow, oh = im.size
    if ow <= 0 or oh <= 0:
        return False
    tw, th = _fit_size(ow, oh, IMG_MAX_SIZE_CM)
    rng = cell.Range
    rng.End = rng.End - 1      # 排除单元格结束符 \x07
    rng.Collapse(0)            # 折叠到单元格末尾内部
    shp = cell.Range.InlineShapes.AddPicture(FileName=path, LinkToFile=False,
                                             SaveWithDocument=True, Range=rng)
    shp.Width = tw
    shp.Height = th
    return True


def _fit_inline_size(shp):
    """将内嵌对象（图片/OLE）等比缩放，使长宽均不超过 IMG_MAX_SIZE_CM。

    实测：图片/Excel 等 OLE 可通过 Width/Height 赋值缩放，
    而 Word 文档类 OLE 图标会忽略 Width/Height 赋值，需改用
    ScaleWidth/ScaleHeight（相对原始尺寸的百分比）。

    注意：对象已在 2cm 内（含浮点尾差）必须直接返回；否则再次调用
    ScaleWidth/ScaleHeight=100 会把已缩好的对象放回原始大小，
    曾实测导致输出文档中图片被放大到自然尺寸、撑爆表格版式。
    """
    max_pt = _cm2pt(IMG_MAX_SIZE_CM)
    try:
        w, h = shp.Width, shp.Height
    except Exception:
        return
    tw, th = _fit_size(w, h, IMG_MAX_SIZE_CM)
    if tw == w and th == h:
        return
    # 方式一：直接设置宽高（对图片/Excel OLE 有效）
    try:
        shp.Width = tw
        shp.Height = th
    except Exception:
        pass
    # 方式二：若宽高赋值未生效（尺寸几乎没变，如 Word 文档类 OLE 图标），
    # 改用百分比缩放（相对原始尺寸），等比缩到 2cm 以内。
    try:
        w2, h2 = shp.Width, shp.Height
    except Exception:
        w2, h2 = w, h
    if w2 <= 0 or h2 <= 0:
        return
    if w2 >= w - 0.5 or h2 >= h - 0.5:
        try:
            scale = min(max_pt / w, max_pt / h)
            shp.ScaleWidth = round(scale * 100, 1)
            shp.ScaleHeight = round(scale * 100, 1)
        except Exception:
            pass


def _insert_ole(cell, path):
    """在单元格末尾插入文件为 OLE 嵌入对象（图标+文件名，可双击打开）。
    保持 Word 自然显示尺寸（图标+文件名标签），随后等比缩小到 IMG_MAX_SIZE_CM 以内。
    """
    name = os.path.basename(path)
    rng = cell.Range
    rng.End = rng.End - 1
    rng.Collapse(0)
    shp = cell.Range.InlineShapes.AddOLEObject(
        FileName=path, LinkToFile=False, DisplayAsIcon=True, IconLabel=name,
        Range=rng)
    _fit_inline_size(shp)
    return True


def _insert_materials(cell, material_paths, tmp_dir=None):
    """向单元格末尾依次插入支撑材料：
    图片 -> 内嵌图片；其他文件 -> OLE 嵌入对象；失败则退化为文件名文字。
    tmp_dir 非空时对超大图片做像素压缩（控制文档体积）。
    """
    valid = [p for p in material_paths if p and os.path.exists(p)]
    for idx, path in enumerate(valid):
        try:
            if is_image(path):
                _insert_image(cell, path, tmp_dir)
            else:
                _insert_ole(cell, path)
        except Exception:
            rng = cell.Range
            rng.End = rng.End - 1
            rng.Collapse(0)
            rng.Text = os.path.basename(path)
        if idx < len(valid) - 1:
            r2 = cell.Range
            r2.End = r2.End - 1
            r2.Collapse(0)
            r2.InsertParagraphAfter()


def _replace_slot(body, label, next_label, value):
    """替换 body 中「label ... next_label」区间内的值，保留区间首尾空白。
    - 若区间原值存在（如旧姓名），新值顶替原值，首尾空白不变；
    - 若区间全为空白（模板留空待填），新值居中放置。
    """
    li = body.find(label)
    if li < 0:
        return body
    start = li + len(label)
    ni = body.find(next_label, start)
    end = ni if ni >= 0 else len(body)
    seg = body[start:end]
    if not seg.strip():
        # 全空白：把空白对半分，值放中间
        n = len(seg)
        head, tail = n // 2, n - n // 2
        return body[:start] + seg[:head] + value + seg[n - tail:] + body[end:]
    head = len(seg) - len(seg.lstrip())
    tail = len(seg) - len(seg.rstrip())
    new_seg = seg[:head] + value + (seg[len(seg) - tail:] if tail else "")
    return body[:start] + new_seg + body[end:]


def _fill_header(cell, name, month):
    """替换表头 R1 中的姓名与评价月份。
    用标签区间截取法，避免正则组引用歧义；保留空白占位区，值居中写入。
    """
    raw = cell.Range.Text
    body = raw.replace("\x07", "").rstrip("\r")

    # 姓名：考评对象（安全员）：...管理者姓名
    body = _replace_slot(body, "考评对象（安全员）：", "管理者姓名", str(name).strip())
    # 月份：评价月份：...评价人员签字
    body = _replace_slot(body, "评价月份：", "评价人员签字", str(month).strip() + "月")

    rng = cell.Range
    rng.End = rng.End - 1
    rng.Text = body


def _sum_scores(items, key):
    """对某字段的数值求和（用于合计），忽略非数值；无有效数值返回空。"""
    total = 0
    has_val = False
    for idx in range(1, 13):
        v = (items.get(idx) or {}).get(key)
        if v is None:
            continue
        try:
            total += float(str(v).strip().rstrip("分"))
            has_val = True
        except (ValueError, TypeError):
            continue
    if not has_val:
        return ""
    return str(int(total)) if total == int(total) else str(total)


def _new_word_app():
    """创建 Word COM 实例。

    必须用静态类型库分派(gencache.EnsureDispatch)：动态分派(Dispatch)下
    InlineShapes.AddOLEObject 在无界面 Word 环境会挂起（实测复现）。
    PyInstaller 打包环境自动回退到动态分派，并优先把类型库缓存指向临时目录。
    """
    if getattr(sys, "frozen", False):
        try:
            gencache.is_readonly = False
            gencache.GetGeneratePath()
        except Exception:
            pass
    try:
        return gencache.EnsureDispatch("Word.Application")
    except Exception:
        return win32com.client.Dispatch("Word.Application")


def _cell_retry(tb, row, col, attempts=2, delay=0.5):
    """读取表格单元格；Word 在嵌入 OLE/大图后偶发表格网格抖动，失败时短延重试。

    两次都失败则照常抛出，由上层统一报错（不产出残缺文档）。
    """
    for i in range(attempts):
        try:
            return tb.Cell(row, col)
        except Exception:
            if i < attempts - 1:
                time.sleep(delay)
    return tb.Cell(row, col)


def generate_doc(template_path, output_path, name, month, items, year=None,
                 progress_cb=None):
    """核心生成：打开模板副本 -> 填充 -> 另存为 .doc。
    items: {1..12: {'desc','score','material_text','material_images':[path],
                    'eval_desc','super_score'}}
    progress_cb: 可选回调，每处理完一个考评项调用 progress_cb(idx)，idx=1..12。
    """
    if year is None:
        year = datetime.now().year
    word = _new_word_app()
    word.Visible = False
    word.DisplayAlerts = 0
    doc = None
    # 图片压缩临时目录：超大图片先压到 MAX_EMBED_PX 内再嵌入，控制 .doc 体积
    tmp_dir = tempfile.mkdtemp(prefix="kaoping_mat_")
    try:
        doc = word.Documents.Open(os.path.abspath(template_path))
        tb = doc.Tables(1)

        # 模板结构校验：防止模板被替换/改版后错格写入
        rows, cols = tb.Rows.Count, tb.Columns.Count
        if rows != TOTAL_ROW or cols != 10:
            raise ValueError(
                f"模板表格结构异常：期望 {TOTAL_ROW} 行×10 列，实际 {rows} 行×{cols} 列，"
                f"请更换正确模板（{TEMPLATE_NAME}）"
            )
        header = tb.Cell(HEADER_ROW, 1).Range.Text
        if "考评对象" not in header or "评价月份" not in header:
            raise ValueError(
                "模板表头缺少「考评对象（安全员）」或「评价月份」标签，请更换正确模板"
            )

        _fill_header(tb.Cell(HEADER_ROW, 1), name, month)

        for idx in range(1, 13):
            row = ITEM_ROWS[idx]
            item = items.get(idx) or {}
            _set_cell_text(_cell_retry(tb, row, COL_SELF_DESC), item.get("desc", ""))
            _set_cell_text(_cell_retry(tb, row, COL_SELF_SCORE), item.get("score", ""))
            mat_cell = _cell_retry(tb, row, COL_MATERIAL)
            _clear_cell_content(mat_cell)
            mat_text = (item.get("material_text") or "").strip()
            materials = [p for p in (item.get("materials")
                                     or item.get("material_images") or [])
                         if p and os.path.exists(p)]
            if mat_text:
                _set_cell_text(mat_cell, mat_text)
            if materials:
                if mat_text:
                    r2 = mat_cell.Range
                    r2.Collapse(0)
                    r2.InsertParagraphAfter()
                _insert_materials(mat_cell, materials, tmp_dir)
            _set_cell_text(_cell_retry(tb, row, COL_EVAL_DESC), item.get("eval_desc", ""))
            _set_cell_text(_cell_retry(tb, row, COL_EVAL_SCORE), item.get("super_score", ""))
            if progress_cb:
                try:
                    progress_cb(idx)
                except Exception:
                    pass

        self_total = _sum_scores(items, "score")
        super_total = _sum_scores(items, "super_score")
        _set_cell_text(tb.Cell(TOTAL_ROW, COL_SELF_DESC), self_total)
        _set_cell_text(tb.Cell(TOTAL_ROW, COL_EVAL_DESC), super_total)

        # 保存前统一缩放 C8 全部内嵌对象（图片/OLE），确保长宽均 ≤ IMG_MAX_SIZE_CM。
        # OLE 对象在插入瞬间尺寸可能未稳定，故在此兜底再缩放一次；
        # 仍有超限者记入日志，便于排查（正常情况下不应再有 2cm 外对象）。
        max_pt = _cm2pt(IMG_MAX_SIZE_CM)
        oversized = []
        for idx in range(1, 13):
            try:
                mat_cell = _cell_retry(tb, ITEM_ROWS[idx], COL_MATERIAL)
                for shp in list(mat_cell.Range.InlineShapes):
                    _fit_inline_size(shp)
                    try:
                        if (shp.Width > max_pt + FIT_TOLERANCE_PT
                                or shp.Height > max_pt + FIT_TOLERANCE_PT):
                            oversized.append((idx, round(shp.Width, 1),
                                              round(shp.Height, 1)))
                    except Exception:
                        pass
            except Exception:
                pass
        if oversized:
            logging.warning("生成文档中仍有超出 2cm 上限的对象: %s", oversized)

        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        # SaveAs2 是 Word 2010+ 的方法；Word 2007 机器自动回退 SaveAs，兼容性更好
        save_as = getattr(doc, "SaveAs2", doc.SaveAs)
        save_as(os.path.abspath(output_path), 0)
    finally:
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        word.Quit()
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return output_path


if __name__ == "__main__":
    tpl = find_template()
    print("模板:", tpl)
    items = {i: {"desc": "", "score": "", "material_text": "",
                 "material_images": [], "eval_desc": "", "super_score": ""}
             for i in range(1, 13)}
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_test_output.doc")
    generate_doc(tpl, out, "测试员", 8, items, year=2026)
    print("已生成:", out, os.path.getsize(out), "bytes")




